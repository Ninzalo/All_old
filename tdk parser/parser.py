import requests
from bs4 import BeautifulSoup
import time
import openpyxl
import os
import random

import undetected_chromedriver.v2 as uc




def create():
    # book = openpyxl.Workbook()
    row = [
        "Name"
        "Length(L)",
        "Width(W)",
        "Thickness(T)",
        "Terminal Width(B)",
        "Terminal Spacing(G)",
        "Recommended Land Pattern (PA)",
        "Recommended Land Pattern (PB)",
        "Recommended Land Pattern (PC)",
        "Capacitance",
        "Rated Voltage",
        "Temperature Characteristic",
        "Dissipation Factor (Max.)",
        "Insulation Resistance (Min.)",
        "Soldering Method",
        "AEC-Q200",
        "Packing",
        "Package Quantity"
    ]
    # for sheet in book.worksheets:
    #     sheet.append(row)


    filename = f"{os.getcwd()}\\test.xlsx"
    book = openpyxl.load_workbook(filename=filename)

    sheet : worksheet = book.worksheets[0]
    sheet.append(row)
    # sheet.insert_rows(0)
    # sheet["A1"].value = "abc"
    # sheet["B1"].value = "cad"

    book.save(filename)


def photo_and_doc(photo_src, doc_src, name):

    try:
        get_doc = requests.get(doc_src)
        with open(f"{os.getcwd()}\\Data\\Docs\\{name}.pdf", "wb") as doc:
            for chunk in get_doc.iter_content(chunk_size=1024):
                if chunk:
                    doc.write(chunk)

    except:
        pass

    try:
        get_photo = requests.get(photo_src)
        with open(f"{os.getcwd()}\\Data\\Photo\\{name}.png", "wb") as photo:
            for chunk in get_photo.iter_content(chunk_size=1024):
                if chunk:
                    photo.write(chunk)
    except:
        pass

def parser():
    global headers
    headers = {
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
        "accept-encoding": "gzip, deflate, br",
        "accept-language": "ru,en;q=0.9,fr;q=0.8",
        "cache-control": "max-age=0",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "Windows",
        "sec-fetch-dest": "document",
        "sec-fetch-mode": "navigate",
        "sec-fetch-site": "none",
        "sec-fetch-user": "?1",
        "upgrade-insecure-requests": "1",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 YaBrowser/21.9.2.169 Yowser/2.5 Safari/537.36"
    }

    product_types = [
         "https://product.tdk.com/en/products/capacitor/index.html",
         "https://product.tdk.com/en/products/inductor/index.html",
         "https://product.tdk.com/en/products/emc/index.html",
         "https://product.tdk.com/en/products/rf/index.html"
                     ]
    catalogs_list = []
    for url in product_types:
        r = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(r.text, "lxml")
        soup = soup.find_all("a", class_="lineup-menus-item")

        for item in soup:
            item = item.get("href")
            if "/catalog" in item:
                # print(f"https://product.tdk.com{item}\n\n\n")
                catalogs_list.append(f"https://product.tdk.com{item}")

        time.sleep(random.randrange(1, 3))
    # print(catalogs_list)
    lists = []

    for url in catalogs_list:
        r = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(r.text, "lxml")
        soup = soup.find_all("td")

        for item in soup:
            item = item.find_all("a")
            for i in item:
                i = i.get("href")
                if "/search" in i:
                    # print(f"https://product.tdk.com{i}\n\n\n")
                    lists.append(f"https://product.tdk.com{i}")

        time.sleep(random.randrange(1, 3))
    print(len(lists))

    with open(f"{os.getcwd()}\\Data\\lists.txt", "w") as f:
        f.write("")

    with open(f"{os.getcwd()}\\Data\\lists.txt", "a") as f:
        for item in lists:
            f.write(f"{item}\n")


    with open(f"{os.getcwd()}\\Data\\lists.txt", "r") as file:
        lists = file.readlines()
        new_list = []
        for item in lists:
            # r = requests.get(url=item, headers=headers)
            # soup = BeautifulSoup(r.text, "lxml")
            item = item.replace("l=20&_p", "l=100&_p")
            new_list.append(item)
        new_list = set(new_list)
        new_list = list(new_list)
        new_list.sort()
        print(len(new_list))

    with open(f"{os.getcwd()}\\Data\\lists.txt", "w") as f:
        f.write("")

    with open(f"{os.getcwd()}\\Data\\lists.txt", "a") as f:
        for item in new_list:
            f.write(f"{item}")




    # with open(f"{os.getcwd()}\\Data\\lists.txt", "r") as file:
    #     lists = file.readlines()
    #     for item in lists[0:1]:

    chrome_options = uc.ChromeOptions()
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-popup-blocking")
    chrome_options.add_argument("--profile-directory=Default")
    chrome_options.add_argument("--ignore-certificate-errors")
    chrome_options.add_argument("--disable-plugins-discovery")
    chrome_options.add_argument("--incognito")
    chrome_options.add_argument("user_agent=DN")
    chrome_options.add_argument("--mute-audio")
    chrome_options.add_argument("--window-size=1032,720")

    driver = uc.Chrome(
        executable_path=r"X:\Programs\Python\tiktok\chromedriver\chromedriver.exe",
        version_main=94,
        options=chrome_options,
    )

    with open(f"{os.getcwd()}\\Data\\lists.txt", "r") as file:
        lists = file.readlines()
        urls = []
        start_iter = 0
        for item in lists[start_iter:-1]:
            print(f"Iter - {start_iter}")
            print(item.replace("\n", ""))
            start_iter += 1
            # multiple_pages = item.split("_p=1")[0] + "_p=XXX" + item.split("_p=1")[-1]
            # print(item.split("_p=1"))
            # print(multiple_pages)
            driver.get(item)
            time.sleep(random.randrange(3, 10))
            # hrefs = driver.find_elements_by_tag_name('a')
            # for product_link in hrefs:
            #     urls.append(product_link.get_attribute("href"))
            # urls = [product_link.get_attribute('href') for product_link in hrefs]
            try:
                pages = driver.find_element_by_xpath("/html/body/div[1]/main/div[2]/div[2]/div[2]/div[1]/div/ul/li[5]/a").text
            except:
                pages = 1
            if pages == 1:
                try:
                    pages = driver.find_element_by_xpath("/html/body/div[1]/main/div[2]/div[2]/div[2]/div[1]/div/ul/li[2]/a").text
                except:
                    pages = 1

            print(f"Pages - {pages}")
            try:
                pages = int(pages)
            except:
                pages = 1
            multiple_pages = []
            for page in range(1, pages+1):
                links_to_pages = item.split("_p=1")[0] + f"_p={str(page)}" + item.split("_p=1")[-1]
                multiple_pages.append(links_to_pages.replace("\n", ""))

            # print(multiple_pages)


            for page in multiple_pages:
                time.sleep(random.randrange(5, 10))
                try:
                    driver.get(page)
                except:
                    pass
                time.sleep(random.randrange(2, 5))
                try:
                    hrefs = driver.find_elements_by_tag_name('a')
                    for product_link in hrefs:

                        product_link = product_link.get_attribute("href")
                        if "/en/search/" in str(product_link) and "/info?" in str(product_link):
                            urls.append(product_link)
                            try:
                                with open(f"{os.getcwd()}\\Data\\urls2.txt", "a") as log_file:
                                    log_file.write(f'{product_link}\n')
                            except:
                                pass
                            # print(product_link.get_attribute("href"))
                except:
                    pass


    counter = 0
    with open(f"{os.getcwd()}\\Data\\urls.txt", "a") as f:
        urls = set(urls)
        urls = list(urls)
        urls.sort()
        for item in urls:
            if "/en/search/" in str(item) and "/info?" in str(item):
                print(item)
                try:
                    f.write(f"{item}\n")
                    counter += 1
                except:
                    pass
    print(counter)

    # with open(f"{os.getcwd()}\\Data\\urls2.txt", "r") as f:
    #     urls2 = f.readlines()
    #     urls2 = set(urls2)
    #     urls2 = list(urls2)
    #     list_of_urls = []
    #     for url in urls2:
    #         list_of_urls.append(url.replace("\n", ""))
    #
    # with open(f"{os.getcwd()}\\Data\\urls2.txt", "w") as f:
    #     f.write(f"{list_of_urls}\n")


def download():

    with open(f"{os.getcwd()}\\Data\\urls.txt", "r") as file:
        urls = file.readlines()
        iteration = 0
        for url in urls:
            r = requests.get(url=url, headers=headers)
            soup = BeautifulSoup(r.text, "lxml")
            try:
                names = soup.find_all("h2", class_="text18 style02")
                for item in names:
                    name = item.text.strip()
                    name = name.split(" ")[0].split("\n")[0]
                    # print(name)
                # print(soup)
            except:
                name = None
            photo_srcs = []
            try:
                list_photo_src = soup.find_all("img")
                for src in list_photo_src:
                    photo_src = src.get("src")
                    if "/product/" in str(photo_src):
                        photo_srcs.append(f"https://product.tdk.com{photo_src}")
                # print(photo_srcs)
            except:
                photo_src = None
            try:
                # list_of_docs = soup.find_all("div", class_="unit_document_area")
                list_of_docs = soup.find_all("a")
                for doc in list_of_docs:
                    doc = doc.get("href")
                    if "chara" in str(doc):
                        doc_src = f"https://product.tdk.com{doc}"
                        # print(doc_src)

                # print(list_of_docs)
            except:
                doc_src = None
            for i in range(0, len(photo_srcs)):
                name = f"{name}_{i}"
                name = name.split("_")[0]
            num_photo = 0
            try:
                for item in photo_srcs:
                    photo_and_doc(photo_src=item, doc_src=None, name=f"{name}_{num_photo}")
                    num_photo += 1
            except:
                pass

            try:
                photo_and_doc(photo_src=None, doc_src=doc_src, name=name)
            except:
                pass

            # print(f"-----{name}-----")

            # try:
            #     length_path = soup.find_all("dt")
            #     row = []
            #     row.append(name)
            #     for item in length_path[6:50]:
            #         if not "Recommended" in str(item):
            #             row.append(item.text.strip())
            #     # print(row)
            #     if "Wave" in str(row) and "Reflow" in str(row):
            #         row2 = []
            #         for item in row:
            #
            #             if "Wave" in str(item):
            #                 row.remove("Wave (Flow)")
            #                 # row2.append(item)
            #             else:
            #                 # row2.append(item)
            #                 pass
            #         for item in row:
            #             if str(item) == "Reflow":
            #                 item = item.replace("Reflow", "Wave (Flow) / Reflow")
            #                 row2.append(item)
            #             else:
            #                 row2.append(item)
            #     else:
            #         row2 = row
            #     print(row2)
            #
            #     # width_path = soup.find_all("dt")
            #     # for item in width_path[7:8]:
            #     #     width = item.text.strip()
            #     # print(width)
            #     #
            #     # Thickness_path = soup.find_all("dt")
            #     # for item in Thickness_path[8:9]:
            #     #     Thickness = item.text.strip()
            #     # print(Thickness)
            #     #
            #     # Terminal_Width_path = soup.find_all("dt")
            #     # for item in Terminal_Width_path[9:10]:
            #     #     Terminal_Width = item.text.strip()
            #     # print(Terminal_Width)
            #     #
            #     # Terminal_Spacing_path = soup.find_all("dt")
            #     # for item in Terminal_Spacing_path[10:11]:
            #     #     Terminal_Spacing = item.text.strip()
            #     # print(Terminal_Spacing)
            #     #
            #     # Recommended_Land_Pattern_path = soup.find_all("dt")
            #     # for item in Recommended_Land_Pattern_path[11:12]:
            #     #     Recommended_Land_Pattern = item.text.strip()
            #     # print(Recommended_Land_Pattern)
            #     #
            #     # Capacitance_path = soup.find_all("dt")
            #     # for item in Capacitance_path[14:15]:
            #     #     Capacitance = item.text.strip()
            #     # print(Capacitance)
            #     #
            #     # Rated_Voltage_path = soup.find_all("dt")
            #     # for item in Rated_Voltage_path[15:16]:
            #     #     Rated_Voltage = item.text.strip()
            #     # print(Rated_Voltage)
            #     #
            #     # Temperature_Characteristic_path = soup.find_all("dt")
            #     # for item in Temperature_Characteristic_path[16:17]:
            #     #     Temperature_Characteristic = item.text.strip()
            #     # print(Temperature_Characteristic)
            #     #
            #     # Dissipation_Factor_path = soup.find_all("dt")
            #     # for item in Dissipation_Factor_path[17:18]:
            #     #     Dissipation_Factor = item.text.strip()
            #     # print(Dissipation_Factor)
            #     #
            #     # Insulation_Resistance_path = soup.find_all("dt")
            #     # for item in Insulation_Resistance_path[18:19]:
            #     #     Insulation_Resistance = item.text.strip()
            #     # print(Insulation_Resistance)
            #     #
            #     # Soldering_Method_path = soup.find_all("dt")
            #     # for item in Soldering_Method_path[19:20]:
            #     #     Soldering_Method = item.text.strip()
            #     # print(Soldering_Method)
            #     #
            #     # AEC_Q200_path = soup.find_all("dt")
            #     # for item in AEC_Q200_path[20:21]:
            #     #     AEC_Q200 = item.text.strip()
            #     # print(AEC_Q200)
            #     #
            #     # Packing_path = soup.find_all("dt")
            #     # for item in Packing_path[21:22]:
            #     #     Packing = item.text.strip()
            #     # print(Packing)
            #     #
            #     # Packing_path = soup.find_all("dt")
            #     # for item in Packing_path[22:23]:
            #     #     Packing = item.text.strip()
            #     # print(Packing)
            #
            #
            #
            #
            # except:
            #     pass
            #
            # filename = f"{os.getcwd()}\\test.xlsx"
            # book = openpyxl.load_workbook(filename=filename)
            #
            # sheet: worksheet = book.worksheets[0]
            # sheet.append(row2)
            # # sheet.insert_rows(0)
            # # sheet["A1"].value = "abc"
            # # sheet["B1"].value = "cad"
            #
            # book.save(filename)
            #
            print(iteration)
            iteration += 1
            time.sleep(random.randrange(1, 3))


# create()
# parser()
download()